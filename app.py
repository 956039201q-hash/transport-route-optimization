"""
智能运输路径优化系统 - Flask 后端
===================================
模块: 订单 / 地址 / 货物 / 容器 / 车辆 / 费用规则 / 路径优化 / 导出
"""

import io
import math
import sqlite3

import requests
from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from vrp_solver import solve_vrp

# ── App setup ────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # #3: 10MB upload limit
CORS(app)

import os
AMAP_WEB_KEY = os.getenv("AMAP_WEB_KEY", "6d352460f8e0acc5ff52b4b07352d185")
DB_PATH = "transport.db"

# ── Database ──────────────────────────────────────────────────────────────────


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")  # #16: Enable FK constraints
    return conn


def _table_columns(conn, table):
    return {r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()}


def _add_col(conn, table, col, coldef):
    cols = _table_columns(conn, table)
    if col not in cols:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {coldef}")


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS _migrations (
            name TEXT PRIMARY KEY,
            applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS orders (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            order_no        TEXT    NOT NULL,
            customer        TEXT,
            address         TEXT    NOT NULL,
            lat             REAL,
            lng             REAL,
            weight          REAL    DEFAULT 0,
            length          REAL    DEFAULT 0,
            width           REAL    DEFAULT 0,
            height          REAL    DEFAULT 0,
            volume          REAL    DEFAULT 0,
            notes           TEXT,
            geocode_status  TEXT    DEFAULT 'pending',
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS vehicles (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    NOT NULL,
            type        TEXT    DEFAULT '中型货车',
            max_weight  REAL    DEFAULT 0,
            max_volume  REAL    DEFAULT 0,
            length      REAL    DEFAULT 0,
            width       REAL    DEFAULT 0,
            height      REAL    DEFAULT 0,
            plate_no    TEXT,
            driver      TEXT,
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS opt_results (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            depot_address  TEXT,
            depot_lat      REAL,
            depot_lng      REAL,
            total_distance REAL,
            created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS route_assignments (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            result_id       INTEGER,
            vehicle_id      INTEGER,
            order_sequence  TEXT,
            total_distance  REAL,
            estimated_time  REAL,
            weight_used     REAL,
            volume_used     REAL
        );

        CREATE TABLE IF NOT EXISTS addresses (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT NOT NULL,
            full_address    TEXT NOT NULL,
            province        TEXT,
            city            TEXT,
            district        TEXT,
            street          TEXT,
            lat             REAL,
            lng             REAL,
            geocode_status  TEXT DEFAULT 'pending',
            unload_rate     REAL DEFAULT 0,
            load_rate       REAL DEFAULT 0,
            default_vehicle_type TEXT DEFAULT '',
            notes           TEXT,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS cargo_items (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            cargo_name      TEXT NOT NULL,
            cargo_code      TEXT,
            weight          REAL DEFAULT 0,
            length          REAL DEFAULT 0,
            width           REAL DEFAULT 0,
            height          REAL DEFAULT 0,
            volume          REAL DEFAULT 0,
            notes           TEXT,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS containers (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT NOT NULL,
            container_type  TEXT DEFAULT '纸箱',
            inner_length    REAL DEFAULT 0,
            inner_width     REAL DEFAULT 0,
            inner_height    REAL DEFAULT 0,
            inner_volume    REAL DEFAULT 0,
            outer_length    REAL DEFAULT 0,
            outer_width     REAL DEFAULT 0,
            outer_height    REAL DEFAULT 0,
            outer_volume    REAL DEFAULT 0,
            stackable       INTEGER DEFAULT 0,
            max_stack_layers INTEGER DEFAULT 1,
            notes           TEXT,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS cargo_container_assoc (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            cargo_id        INTEGER NOT NULL,
            container_id    INTEGER NOT NULL,
            items_per_container INTEGER DEFAULT 1,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS region_vehicle_rules (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            province        TEXT NOT NULL,
            vehicle_type    TEXT NOT NULL,
            vehicle_id      INTEGER,
            notes           TEXT,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS order_batches (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT NOT NULL,
            file_name       TEXT,
            order_count     INTEGER DEFAULT 0,
            total_weight    REAL DEFAULT 0,
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS vehicle_cost_rules (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_type       TEXT NOT NULL,
            calc_method        TEXT DEFAULT 'per_mileage',
            mileage_unit_price REAL DEFAULT 0,
            base_price         REAL DEFAULT 0,
            extra_stop_fee     REAL DEFAULT 0,
            misc_fees          REAL DEFAULT 0,
            daily_rate         REAL DEFAULT 0,
            hourly_rate        REAL DEFAULT 0,
            volume_rate        REAL DEFAULT 0,
            weight_rate        REAL DEFAULT 0,
            notes              TEXT,
            created_at         TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    _add_col(conn, "vehicles", "loading_method", "TEXT DEFAULT '后卸车'")
    _add_col(conn, "vehicles", "weight_load_rate", "REAL DEFAULT 70")
    _add_col(conn, "vehicles", "volume_load_rate", "REAL DEFAULT 70")

    _add_col(conn, "orders", "address_id", "INTEGER")
    _add_col(conn, "orders", "cargo_id", "INTEGER")
    _add_col(conn, "orders", "container_id", "INTEGER")
    _add_col(conn, "orders", "quantity", "INTEGER DEFAULT 1")
    _add_col(conn, "orders", "order_date", "DATE")
    _add_col(conn, "orders", "delivery_date", "DATE")
    _add_col(conn, "orders", "receipt_date", "DATE")
    _add_col(conn, "orders", "batch_id", "INTEGER")

    _add_col(conn, "vehicle_cost_rules", "vehicle_id", "INTEGER")
    _add_col(conn, "vehicle_cost_rules", "vehicle_name", "TEXT")

    _add_col(conn, "route_assignments", "transport_cost", "REAL DEFAULT 0")
    _add_col(conn, "route_assignments", "loading_rate", "REAL DEFAULT 0")
    _add_col(conn, "route_assignments", "is_return_trip", "INTEGER DEFAULT 1")

    _add_col(conn, "opt_results", "return_trip", "INTEGER DEFAULT 1")

    _add_col(conn, "orders", "cargo_code", "TEXT DEFAULT ''")
    _add_col(conn, "orders", "piece_count", "INTEGER DEFAULT 0")
    _add_col(conn, "orders", "pallet_count", "INTEGER DEFAULT 1")
    _add_col(conn, "orders", "pallet_type", "TEXT DEFAULT ''")
    _add_col(conn, "orders", "stack_layers", "INTEGER DEFAULT 1")

    _add_col(conn, "orders", "is_deleted", "INTEGER DEFAULT 0")
    _add_col(conn, "orders", "deleted_at", "TIMESTAMP")
    _add_col(conn, "opt_results", "name", "TEXT DEFAULT ''")
    _add_col(conn, "route_assignments", "delivery_status", "TEXT DEFAULT 'pending'")

    done = conn.execute("SELECT 1 FROM _migrations WHERE name='cm_to_mm_v1'").fetchone()
    if not done:
        conn.execute("UPDATE vehicles SET length=length*10, width=width*10, height=height*10 WHERE length>0 AND length<5000")
        conn.execute("UPDATE orders SET length=length*10, width=width*10, height=height*10 WHERE length>0 AND length<5000")
        conn.execute("INSERT INTO _migrations (name) VALUES ('cm_to_mm_v1')")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            action     TEXT NOT NULL,
            entity     TEXT,
            entity_id  INTEGER,
            detail     TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.executescript("""
        CREATE INDEX IF NOT EXISTS idx_orders_batch_id ON orders(batch_id);
        CREATE INDEX IF NOT EXISTS idx_orders_order_date ON orders(order_date);
        CREATE INDEX IF NOT EXISTS idx_orders_is_deleted ON orders(is_deleted);
        CREATE INDEX IF NOT EXISTS idx_orders_geocode ON orders(geocode_status);
        CREATE INDEX IF NOT EXISTS idx_route_assignments_result_id ON route_assignments(result_id);
        CREATE INDEX IF NOT EXISTS idx_audit_log_created ON audit_log(created_at);
    """)

    conn.commit()
    conn.close()


def log_action(conn, action, entity=None, entity_id=None, detail=None):
    conn.execute(
        "INSERT INTO audit_log (action, entity, entity_id, detail) VALUES (?,?,?,?)",
        (action, entity, entity_id, detail),
    )


def haversine(lat1, lon1, lat2, lon2):
    R = 6_371_000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def geocode_address(address: str):
    try:
        resp = requests.get(
            "https://restapi.amap.com/v3/geocode/geo",
            params={"address": address, "key": AMAP_WEB_KEY, "output": "json"},
            timeout=10,
        )
        data = resp.json()
        if data.get("status") == "1" and data.get("geocodes"):
            gc = data["geocodes"][0]
            loc = gc["location"]
            lng, lat = map(float, loc.split(","))
            province = gc.get("province", "")
            city = gc.get("city", "") or ""
            district = gc.get("district", "") or ""
            street = gc.get("street", "") or ""
            if isinstance(province, list): province = province[0] if province else ""
            if isinstance(city, list): city = city[0] if city else ""
            return lat, lng, "success", province, city, district, street
        return None, None, "failed", "", "", "", ""
    except Exception as exc:
        return None, None, f"error:{exc}", "", "", "", ""


def _style_header(ws, row_num, fill_hex="1e3a5f"):
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(fill_type="solid", fgColor=fill_hex)
    halign = Alignment(horizontal="center", vertical="center")
    for cell in ws[row_num]:
        cell.font = hf
        cell.fill = hfill
        cell.alignment = halign


def _auto_col_width(ws, max_width=45):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, max_width)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/test", methods=["GET"])
def test():
    return jsonify({"status": "ok", "message": "API is working"})


if __name__ == "__main__":
    init_db()
    print("\n[OK] SmartRoute started")
    print("URL: http://localhost:5000\n")
    app.run(debug=True, port=5000, host="0.0.0.0")
